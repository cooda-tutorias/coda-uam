from django.test import TestCase
from django.test import override_settings
from django.db import models
from django.core.files.uploadedfile import SimpleUploadedFile
from Usuarios.models import Alumno, Tutor, Cordinador, Coda, Documento
from Usuarios.constants import ALUMNO, TUTOR, COORDINADOR, CODA
from django.contrib.auth import get_user_model
from django.urls import reverse
import json
import tempfile
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from tablib import Dataset
from webpush.models import PushInformation, SubscriptionInfo
from Usuarios.admin import TutorAdmin, TutorResource
from Usuarios.models import PushDevice
from Usuarios.services.importacion_alumnos import (
    ENCABEZADOS_IMPORTACION,
    ENCABEZADOS_OPCIONALES,
    importar_alumnos_validados,
    leer_y_normalizar_alumnos,
    normalizar_fila_alumno,
    obtener_trimestre_ingreso,
    validar_archivo_alumnos,
)
from Usuarios.services.plantilla_importacion_alumnos import (
    generar_plantilla_importacion_alumnos,
)
from Usuarios.services.importacion_tutores import (
    validar_y_normalizar_dataset_tutores,
)

# Importar el modelo de Usuario
Usuario = get_user_model()


class TutorResourceTests(TestCase):
    encabezados = [
        "matricula",
        "first_name",
        "last_name",
        "second_last_name",
        "email",
        "sexo",
        "coordinacion",
        "cubiculo",
        "password",
    ]

    def dataset(self, *filas):
        dataset = Dataset(headers=self.encabezados)
        for fila in filas:
            dataset.append(fila)
        return dataset

    def fila_valida(self, **cambios):
        datos = {
            "matricula": "30419",
            "first_name": "Antonio",
            "last_name": "López",
            "second_last_name": "",
            "email": "alopez.importacion@cua.uam.mx",
            "sexo": "M",
            "coordinacion": "COM",
            "cubiculo": "723B",
            "password": "Temporal-12345",
        }
        datos.update(cambios)
        return [datos[encabezado] for encabezado in self.encabezados]

    def test_tutor_admin_utiliza_recurso_especifico(self):
        self.assertIs(TutorAdmin.resource_class, TutorResource)
        self.assertEqual(TutorResource._meta.import_id_fields, ("matricula",))
        self.assertNotIn("id", TutorResource._meta.fields)

    def test_importa_campos_necesarios_y_hashea_password(self):
        dataset = self.dataset(self.fila_valida(
            matricula=30419,
            first_name="  Antonio ",
            email=" ALOPEZ.IMPORTACION@CUA.UAM.MX ",
            sexo=" masculino ",
            coordinacion=" ingeniería en computación ",
            cubiculo=" 723b ",
        ))

        resultado = TutorResource().import_data(
            dataset,
            dry_run=False,
            raise_errors=True,
        )

        self.assertFalse(resultado.has_errors())
        tutor = Tutor.objects.get(matricula="30419")
        self.assertEqual(tutor.first_name, "Antonio")
        self.assertEqual(tutor.last_name, "López")
        self.assertEqual(tutor.second_last_name, "")
        self.assertEqual(tutor.sexo, "M")
        self.assertEqual(tutor.coordinacion, "COM")
        self.assertEqual(tutor.cubiculo, "723B")
        self.assertIn(TUTOR, tutor.rol)
        self.assertTrue(tutor.es_tutor)
        self.assertFalse(tutor.es_coordinador)
        self.assertTrue(tutor.check_password("Temporal-12345"))
        self.assertNotEqual(tutor.password, "Temporal-12345")

    def test_rechaza_tutor_existente_sin_actualizarlo(self):
        existente = Tutor.objects.create_user(
            matricula="30419",
            email="existente@cua.uam.mx",
            password="Original-12345",
            rol=[TUTOR],
            coordinacion="MAT",
            cubiculo="601",
        )
        dataset = self.dataset(self.fila_valida())

        resultado = TutorResource().import_data(dataset, dry_run=False)

        self.assertTrue(resultado.has_errors())
        existente.refresh_from_db()
        self.assertEqual(existente.email, "existente@cua.uam.mx")
        self.assertEqual(existente.coordinacion, "MAT")
        self.assertTrue(existente.check_password("Original-12345"))

    def test_duplicados_cancelan_archivo_completo(self):
        dataset = self.dataset(
            self.fila_valida(),
            self.fila_valida(first_name="Otra persona"),
        )

        resultado = TutorResource().import_data(dataset, dry_run=False)

        self.assertTrue(resultado.has_errors())
        self.assertFalse(Tutor.objects.filter(matricula="30419").exists())

    def test_fila_invalida_cancela_tambien_las_filas_validas(self):
        dataset = self.dataset(
            self.fila_valida(),
            self.fila_valida(
                matricula="30A20",
                email="correo-invalido",
                sexo="Otro",
                coordinacion="Carrera inexistente",
                cubiculo="7-23",
                password="123",
            ),
        )

        resultado = TutorResource().import_data(dataset, dry_run=False)

        self.assertTrue(resultado.has_errors())
        self.assertEqual(Tutor.objects.count(), 0)

    def test_servicio_reporta_todos_los_campos_invalidos(self):
        dataset = self.dataset(self.fila_valida(
            matricula="30A19",
            first_name="",
            last_name="",
            second_last_name="",
            email="correo-invalido",
            sexo="Otro",
            coordinacion="Carrera inexistente",
            cubiculo="7-23",
            password="123",
        ))

        resultado = validar_y_normalizar_dataset_tutores(dataset)
        campos = {error.campo for error in resultado.errores}

        self.assertFalse(resultado.es_valido)
        self.assertTrue({
            "matricula", "first_name", "last_name",
            "email", "sexo", "coordinacion", "cubiculo", "password",
        }.issubset(campos))
        self.assertNotIn("second_last_name", campos)


class NormalizacionImportacionAlumnosTests(TestCase):
    def test_normaliza_catalogos_correos_y_espacios(self):
        fila = normalizar_fila_alumno({
            "Plan de estudios": "  ingeniería EN computación ",
            "Matrícula": "2223028388",
            "Correo institucional": " ALUMNO@CUA.UAM.MX ",
            "Correo alterno": " PERSONAL@EXAMPLE.COM ",
            "Apellido Paterno": "  López  ",
            "Apellido Materno": "  García  ",
            "Nombres": "  María Fernanda  ",
            "Núm. económico tutor": 12345,
            "Estado académico": 10,
            "Sexo": " fEmEnInO ",
        }, numero_fila=7)

        self.assertEqual(fila["numero_fila"], 7)
        self.assertEqual(fila["carrera"], "COM")
        self.assertEqual(fila["carrera_nombre"], "Ingeniería en Computación")
        self.assertEqual(fila["sexo"], "F")
        self.assertEqual(fila["sexo_nombre"], "Femenino")
        self.assertEqual(fila["email"], "alumno@cua.uam.mx")
        self.assertEqual(fila["correo_personal"], "personal@example.com")
        self.assertEqual(fila["first_name"], "María Fernanda")
        self.assertEqual(fila["last_name"], "López")
        self.assertEqual(fila["second_last_name"], "García")
        self.assertEqual(fila["tutor_matricula"], "12345")
        self.assertEqual(fila["estado"], "10")

    def test_deriva_trimestre_de_matriculas_de_nueve_y_diez_digitos(self):
        self.assertEqual(obtener_trimestre_ingreso("209373965"), "09-O")
        self.assertEqual(obtener_trimestre_ingreso("2223028388"), "22-O")

    def test_conserva_si_la_matricula_original_era_texto(self):
        texto = normalizar_fila_alumno({"Matrícula": "2223028388"}, 2)
        numero = normalizar_fila_alumno({"Matrícula": 2223028388}, 3)

        self.assertTrue(texto["matricula_original_es_texto"])
        self.assertFalse(numero["matricula_original_es_texto"])
        self.assertFalse(texto["matricula_original_es_entero"])
        self.assertTrue(numero["matricula_original_es_entero"])

    def test_lee_csv_y_reporta_numero_real_de_fila(self):
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                "Plan de estudios,Matrícula,Correo institucional,Correo alterno,"
                "Apellido Paterno,Apellido Materno,Nombres,Núm. económico tutor,Estado académico,Sexo\n"
                "Biología Molecular,2223028388,A@CUA.UAM.MX,p@example.com,"
                "López,García,Ana,12345,1,FEMENINO\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        filas = leer_y_normalizar_alumnos(archivo)

        self.assertEqual(len(filas), 1)
        self.assertEqual(filas[0]["numero_fila"], 2)
        self.assertEqual(filas[0]["carrera"], "BM")
        self.assertEqual(filas[0]["sexo"], "F")
        self.assertEqual(filas[0]["trimestre_ingreso"], "22-O")


class ValidacionImportacionAlumnosTests(TestCase):
    encabezado = (
        "Plan de estudios,Matrícula,Correo institucional,Correo alterno,"
        "Apellido Paterno,Apellido Materno,Nombres,Núm. económico tutor,Estado académico,Sexo\n"
    )

    def setUp(self):
        self.tutor = Tutor.objects.create_user(
            matricula="12345",
            email="tutor.importacion@example.com",
            password="password123",
            rol=[TUTOR],
            coordinacion="COM",
            first_name="José Antonio",
            last_name="Pérez",
            second_last_name="Hernández",
        )

    def archivo(self, *filas):
        return SimpleUploadedFile(
            "alumnos.csv",
            (self.encabezado + "".join(filas)).encode("utf-8"),
            content_type="text/csv",
        )

    def test_acepta_varios_correos_personales_vacios(self):
        archivo = self.archivo(
            "Ingeniería en Computación,2223028388,a1@cua.uam.mx,,López,,Ana,12345,1,Femenino\n",
            "biología molecular,2232028389,a2@cua.uam.mx,,Pérez,Gómez,Luis,12345,19,masculino\n",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertTrue(resultado.es_valido)
        self.assertEqual(resultado.total_filas, 2)
        self.assertEqual(resultado.errores, [])
        self.assertEqual(resultado.advertencias, [])

    def test_acepta_matricula_entera_y_rechaza_decimal_y_notacion_cientifica(self):
        libro = Workbook()
        hoja = libro.active
        encabezados = ENCABEZADOS_IMPORTACION + ENCABEZADOS_OPCIONALES
        hoja.append(encabezados)
        base = {
            "Plan de estudios": "Biología Molecular",
            "Apellido Paterno": "López",
            "Apellido Materno": "",
            "Nombres": "Ana",
            "Sexo": "Femenino",
            "Estado académico": 1,
            "Correo alterno": "",
            "Núm. económico tutor": 12345,
            "Nombre del tutor": "",
        }
        casos = (
            (2223028388, "entero@cua.uam.mx"),
            (2232028389.5, "decimal@cua.uam.mx"),
            ("2.243076344E9", "cientifica@cua.uam.mx"),
        )
        for matricula, correo in casos:
            fila = base.copy()
            fila["Matrícula"] = matricula
            fila["Correo institucional"] = correo
            hoja.append([fila.get(encabezado, "") for encabezado in encabezados])

        contenido = BytesIO()
        libro.save(contenido)
        archivo = SimpleUploadedFile(
            "alumnos.xlsx",
            contenido.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        resultado = validar_archivo_alumnos(archivo)
        errores_matricula = [
            error for error in resultado.errores
            if error.columna == "Matrícula"
        ]

        self.assertNotIn(2, {error.fila for error in errores_matricula})
        self.assertIn(3, {error.fila for error in errores_matricula})
        self.assertIn(4, {error.fila for error in errores_matricula})

    def test_nombre_parcial_y_sin_acentos_coincide_con_tutor(self):
        encabezado = self.encabezado.rstrip("\n") + ",Nombre del tutor\n"
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                encabezado
                + "Ingeniería en Computación,2223028388,a1@cua.uam.mx,,"
                  "López,,Ana,12345,1,Femenino,Jose Perez\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertTrue(resultado.es_valido)
        self.assertEqual(resultado.advertencias, [])

    def test_nombre_con_menos_de_sesenta_por_ciento_genera_advertencia(self):
        encabezado = self.encabezado.rstrip("\n") + ",Nombre del tutor\n"
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                encabezado
                + "Ingeniería en Computación,2223028388,a1@cua.uam.mx,,"
                  "López,,Ana,12345,1,Femenino,María López\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertTrue(resultado.es_valido)
        self.assertEqual(len(resultado.advertencias), 1)
        self.assertEqual(resultado.advertencias[0].fila, 2)
        self.assertEqual(resultado.advertencias[0].columna, "Nombre del tutor")
        self.assertIn("José Antonio Pérez Hernández", resultado.advertencias[0].mensaje)

    def test_nombre_de_tutor_vacio_no_genera_advertencia(self):
        encabezado = self.encabezado.rstrip("\n") + ",Nombre del tutor\n"
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                encabezado
                + "Ingeniería en Computación,2223028388,a1@cua.uam.mx,,"
                  "López,,Ana,12345,1,Femenino,\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertTrue(resultado.es_valido)
        self.assertEqual(resultado.advertencias, [])

    def test_tutor_inexistente_conserva_error_sin_comparar_nombre(self):
        encabezado = self.encabezado.rstrip("\n") + ",Nombre del tutor\n"
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                encabezado
                + "Ingeniería en Computación,2223028388,a1@cua.uam.mx,,"
                  "López,,Ana,99999,1,Femenino,María López\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertFalse(resultado.es_valido)
        self.assertTrue(any(
            error.columna == "Núm. económico tutor"
            and "registrado" in error.mensaje
            for error in resultado.errores
        ))
        self.assertEqual(resultado.advertencias, [])

    def test_reporta_duplicado_solo_si_correo_personal_tiene_valor(self):
        archivo = self.archivo(
            "Ingeniería en Computación,2223028388,a1@cua.uam.mx,repetido@example.com,López,,Ana,12345,1,Femenino\n",
            "Biología Molecular,2232028389,a2@cua.uam.mx,REPETIDO@example.com,Pérez,,Luis,12345,2,Masculino\n",
        )

        resultado = validar_archivo_alumnos(archivo)
        errores_correo = [
            error for error in resultado.errores
            if error.columna == "Correo alterno" and "repetido" in error.mensaje
        ]

        self.assertFalse(resultado.es_valido)
        self.assertEqual({error.fila for error in errores_correo}, {2, 3})

    def test_reporta_correos_mal_formados_y_dominio_gmai(self):
        archivo = self.archivo(
            "Ingeniería en Computación,2223028388,a1@cua.uam.mx,tonio.jaimes@gmai.com,López,,Ana,12345,1,Femenino\n",
            "Biología Molecular,2232028389,a2@cua.uam.mx,tonio.jaimes@hotmail..com,Pérez,,Luis,12345,2,Masculino\n",
        )

        resultado = validar_archivo_alumnos(archivo)
        errores_correo = [
            error for error in resultado.errores
            if error.columna == "Correo alterno"
        ]

        self.assertEqual({error.fila for error in errores_correo}, {2, 3})
        self.assertTrue(any("gmail.com" in error.mensaje for error in errores_correo))

    def test_reune_errores_de_distintas_filas_sin_escribir(self):
        usuarios_antes = Usuario.objects.count()
        archivo = self.archivo(
            "Carrera inexistente,2224028388,correo-invalido,,López,,,99999,20,Otro\n",
            "Biología Molecular,2232028389,a2@cua.uam.mx,,Pérez,,Luis,12345,1,Femenino\n",
        )

        resultado = validar_archivo_alumnos(archivo)

        self.assertFalse(resultado.es_valido)
        self.assertEqual({error.fila for error in resultado.errores}, {2})
        self.assertGreaterEqual(len(resultado.errores), 6)
        self.assertEqual(Usuario.objects.count(), usuarios_antes)


class ImportAlumnosViewTests(TestCase):
    encabezado = (
        "Plan de estudios,Matrícula,Correo institucional,Correo alterno,"
        "Apellido Paterno,Apellido Materno,Nombres,Núm. económico tutor,Estado académico,Sexo\n"
    )

    def setUp(self):
        self.coda = Coda.objects.create_user(
            matricula="CODIMPORT",
            email="coda.importacion@example.com",
            password="password123",
            rol=[CODA],
        )
        self.tutor = Tutor.objects.create_user(
            matricula="12345",
            email="tutor.fase2@example.com",
            password="password123",
            rol=[TUTOR],
            coordinacion="COM",
            first_name="Antonio",
            last_name="López",
            second_last_name="Jaimes",
        )
        self.client.force_login(self.coda)
        session = self.client.session
        session["role"] = "coda"
        session.save()

    def archivo(self, fila):
        return SimpleUploadedFile(
            "alumnos.csv",
            (self.encabezado + fila).encode("utf-8"),
            content_type="text/csv",
        )

    def test_archivo_invalido_muestra_fila_y_no_crea_alumno(self):
        archivo = self.archivo(
            "Biología Molecular,2223028388,correo-malo,,López,,Ana,12345,1,Femenino\n"
        )

        response = self.client.post(reverse("importar-alumnos"), {"archivo": archivo})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se guardó ningún alumno")
        self.assertContains(response, "Correo institucional")
        self.assertContains(response, "<td>2</td>", html=True)
        self.assertContains(response, "correo-malo")
        self.assertContains(response, "Vista previa")
        self.assertFalse(Alumno.objects.filter(matricula="2223028388").exists())

    def test_archivo_valido_crea_alumno_con_datos_normalizados(self):
        archivo = self.archivo(
            "biología molecular,2223028388,A1@CUA.UAM.MX,, López , , Ana ,12345,1,femenino\n"
        )

        response = self.client.post(reverse("importar-alumnos"), {"archivo": archivo})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "1 alumno importado correctamente")
        alumno = Alumno.objects.get(matricula="2223028388")
        self.assertEqual(alumno.email, "a1@cua.uam.mx")
        self.assertIsNone(alumno.correo_personal)
        self.assertEqual(alumno.first_name, "Ana")
        self.assertEqual(alumno.last_name, "López")
        self.assertIsNone(alumno.second_last_name)
        self.assertEqual(alumno.carrera, "BM")
        self.assertEqual(alumno.estado, 1)
        self.assertEqual(alumno.sexo, "F")
        self.assertEqual(alumno.trimestre_ingreso, "22-O")
        self.assertEqual(alumno.tutor_asignado, self.tutor)
        self.assertEqual(alumno.rol, [ALUMNO])
        self.assertTrue(alumno.check_password("2223028388"))
        self.assertNotEqual(alumno.password, "2223028388")

    def test_advertencia_por_nombre_de_tutor_no_impide_importacion(self):
        encabezado = self.encabezado.rstrip("\n") + ",Nombre del tutor\n"
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                encabezado
                + "Biología Molecular,2223028388,a1@cua.uam.mx,,López,,"
                  "Ana,12345,1,Femenino,Persona Diferente\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(reverse("importar-alumnos"), {"archivo": archivo})

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "El nombre «Persona Diferente» parece no corresponder",
        )
        self.assertContains(response, "Antonio López Jaimes")
        self.assertContains(response, "Persona Diferente")
        self.assertContains(response, "1 alumno importado correctamente")
        self.assertTrue(Alumno.objects.filter(matricula="2223028388").exists())

    def test_fallo_intermedio_revierte_todo_el_lote(self):
        archivo = SimpleUploadedFile(
            "alumnos.csv",
            (
                self.encabezado
                + "Biología Molecular,2223028388,a1@cua.uam.mx,,López,,Ana,12345,1,Femenino\n"
                + "Ingeniería en Computación,2232028389,a2@cua.uam.mx,,Pérez,,Luis,12345,2,Masculino\n"
            ).encode("utf-8"),
            content_type="text/csv",
        )
        resultado = validar_archivo_alumnos(archivo)
        self.assertTrue(resultado.es_valido)
        crear_original = Alumno.objects.create_user
        llamadas = 0

        def crear_y_fallar(**datos):
            nonlocal llamadas
            llamadas += 1
            if llamadas == 2:
                raise RuntimeError("fallo simulado")
            return crear_original(**datos)

        with patch.object(Alumno.objects, "create_user", side_effect=crear_y_fallar):
            with self.assertRaises(RuntimeError):
                importar_alumnos_validados(resultado.filas_validas)

        self.assertFalse(Alumno.objects.filter(
            matricula__in=["2223028388", "2232028389"],
        ).exists())


class PlantillaImportacionAlumnosTests(TestCase):
    def setUp(self):
        self.coda = Coda.objects.create_user(
            matricula="CODTPL001",
            email="coda.plantilla@example.com",
            password="password123",
            rol=[CODA],
        )
        self.tutor = Tutor.objects.create_user(
            matricula="TUTTPL001",
            email="tutor.plantilla@example.com",
            password="password123",
            rol=[TUTOR],
            coordinacion="COM",
        )
        self.url = reverse("plantilla-importacion-alumnos")

    def test_generador_usa_encabezados_actuales_y_formato_de_matricula(self):
        contenido = generar_plantilla_importacion_alumnos()
        libro = load_workbook(BytesIO(contenido))

        self.assertEqual(libro.sheetnames, ["Alumnos", "Instrucciones", "Catálogos"])
        hoja = libro["Alumnos"]
        encabezados = tuple(
            hoja.cell(row=1, column=columna).value
            for columna in range(1, hoja.max_column + 1)
        )
        self.assertEqual(
            encabezados,
            ENCABEZADOS_IMPORTACION + ENCABEZADOS_OPCIONALES,
        )
        columna_matricula = encabezados.index("Matrícula") + 1
        self.assertEqual(hoja.cell(row=2, column=columna_matricula).number_format, "@")
        self.assertEqual(len(hoja.data_validations.dataValidation), 3)
        self.assertTrue(all(
            hoja.cell(row=2, column=columna).value is None
            for columna in range(1, hoja.max_column + 1)
        ))

    def test_coda_descarga_archivo_xlsx_valido(self):
        self.client.force_login(self.coda)
        session = self.client.session
        session["role"] = "coda"
        session.save()

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("plantilla_importacion_alumnos.xlsx", response["Content-Disposition"])
        libro = load_workbook(BytesIO(response.content))
        self.assertIn("Alumnos", libro.sheetnames)

    def test_usuario_anonimo_no_puede_descargar_plantilla(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertIn("next=", response.url)

    def test_tutor_no_puede_descargar_plantilla(self):
        self.client.force_login(self.tutor)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)


class EliminarDocumentoAccessTests(TestCase):
    def setUp(self):
        self.media_directory = tempfile.TemporaryDirectory()
        self.settings_override = override_settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(self.media_directory.cleanup)

        self.password = "password123"
        self.coda = Coda.objects.create_user(
            email="coda.documentos@example.com",
            matricula="CODADOC001",
            password=self.password,
            rol=[CODA],
        )
        self.tutor = Tutor.objects.create_user(
            email="tutor.documentos@example.com",
            matricula="TUTDOC001",
            password=self.password,
            rol=[TUTOR],
            coordinacion="COM",
        )
        self.documento = Documento.objects.create(
            nombre="Plantilla protegida",
            archivo=SimpleUploadedFile("plantilla.txt", b"contenido"),
        )
        self.url = reverse("eliminar_documento", args=[self.documento.pk])

    def test_usuario_anonimo_no_puede_eliminar_documento(self):
        response = self.client.post(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Documento.objects.filter(pk=self.documento.pk).exists())

    def test_usuario_sin_rol_coda_no_puede_eliminar_documento(self):
        self.client.force_login(self.tutor)

        response = self.client.post(self.url)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Documento.objects.filter(pk=self.documento.pk).exists())

    def test_get_no_elimina_documento(self):
        self.client.force_login(self.coda)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 405)
        self.assertTrue(Documento.objects.filter(pk=self.documento.pk).exists())

    def test_coda_puede_eliminar_documento_mediante_post(self):
        archivo = self.documento.archivo
        self.assertTrue(archivo.storage.exists(archivo.name))
        self.client.force_login(self.coda)

        response = self.client.post(self.url)

        self.assertRedirects(response, reverse("ajustes"))
        self.assertFalse(Documento.objects.filter(pk=self.documento.pk).exists())
        self.assertFalse(archivo.storage.exists(archivo.name))


class VerAlumnosCODDAAAccessTests(TestCase):
    def setUp(self):
        self.password = "password123"
        self.coda = Coda.objects.create_user(
            email="coda.alumnos@example.com",
            matricula="CODALUM001",
            password=self.password,
            rol=[CODA],
        )
        self.tutor = Tutor.objects.create_user(
            email="tutor.alumnos@example.com",
            matricula="TUTALUM001",
            password=self.password,
            rol=[TUTOR],
            coordinacion="COM",
        )
        self.alumno = Alumno.objects.create_user(
            email="alumno.listado@example.com",
            matricula="ALULIST001",
            password=self.password,
            rol=[ALUMNO],
            carrera="COM",
            tutor_asignado=self.tutor,
        )
        self.url = reverse("ver-alumnos")

    def test_usuario_anonimo_no_puede_ver_listado_de_alumnos(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertIn("next=", response.url)

    def test_usuario_sin_rol_coda_no_puede_ver_listado_de_alumnos(self):
        self.client.force_login(self.tutor)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_coda_puede_ver_listado_de_alumnos(self):
        self.client.force_login(self.coda)
        session = self.client.session
        session["role"] = "coda"
        session.save()

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertIn(self.alumno, response.context["alumnos"])

# Clase para probar que el campo "cubiculo" es un CharField en los 
# modelos Tutor, Cordinador y Coda
class CubiculoTestCase(TestCase):

    def test_campo_cubiculo_es_charfield(self):
        modelos = [Tutor, Cordinador, Coda]
        
        for modelo in modelos:
            campo = modelo._meta.get_field('cubiculo')
            # Verifica que el campo sea de tipo models.CharField
            self.assertIsInstance(campo, models.CharField)

            
    def test_tipo_interno_cubiculo(self):
            modelos = [Tutor, Cordinador, Coda]
            
            for modelo in modelos:
                tipo = modelo._meta.get_field('cubiculo').get_internal_type()
                # Compara la cadena de texto 'CharField'
                self.assertEqual(tipo, 'CharField')


# Clase para probar que el nombre completo de los usuarios se muestra correctamente
# aún cuando no tienen apellido materno
class UsuarioNombreCompletoTest(TestCase):
    def test_nombre_completo_con_apellido_materno(self):
        # Crear un usuario con apellido materno
        usuario = Usuario(
            first_name="Antonio",
            last_name="López",
            second_last_name="Jaimes"
        )
        self.assertEqual(usuario.nombre_completo, "Antonio López Jaimes")

    def test_nombre_completo_sin_apellido_materno(self):
        # Crear un usuario sin apellido materno
        usuario = Usuario(
            first_name="Mika",
            last_name="Olsen",
            second_last_name=None  # No tiene apellido materno
        )

        #1. Comprobamos que el nombre completo se genera correctamente sin el apellido materno.
        self.assertEqual(usuario.nombre_completo, "Mika Olsen")

        #2. Reforzamos que el apellido materno es None para este usuario.
        self.assertNotIn("None", usuario.nombre_completo)


# Clase para probar que la página de login preserva la URL de redirección
# cuando un alumno escanea el QR de un código de tutoría in situ y es 
# redirigido a la página de login.
class LoginRedirectTest(TestCase):
    def test_login_page_preserves_next_url(self):
        response = self.client.get(reverse("login"), {"next": "/tutorias/in-situ/7/"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="next"')
        self.assertContains(response, 'value="/tutorias/in-situ/7/"')


class LoginSinSelectorDeRolTests(TestCase):
    def setUp(self):
        self.password = "password123"
        self.tutor = Tutor.objects.create_user(
            email="login.tutor@example.com",
            matricula="LOGIN001",
            password=self.password,
            rol=[TUTOR],
            coordinacion="COM",
        )
        self.alumno = Alumno.objects.create_user(
            email="login.alumno@example.com",
            matricula="LOGIN002",
            password=self.password,
            rol=[ALUMNO],
            carrera="COM",
            tutor_asignado=self.tutor,
        )

    def test_login_no_muestra_selector_de_rol(self):
        response = self.client.get(reverse("login"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="role"')
        self.assertNotContains(response, "Seleccionar Rol")

    def test_login_infiere_unico_rol_abre_panel_y_lo_guarda_en_sesion(self):
        response = self.client.post(reverse("login"), {
            "username": self.tutor.email,
            "password": self.password,
        })

        self.assertRedirects(
            response,
            reverse("Panel-tutorias-tutor"),
            fetch_redirect_response=False,
        )
        self.assertEqual(self.client.session["role"], "tutor")
        self.assertEqual(int(self.client.session["_auth_user_id"]), self.tutor.pk)

    def test_login_de_alumno_abre_su_panel_de_tutorias(self):
        response = self.client.post(reverse("login"), {
            "username": self.alumno.email,
            "password": self.password,
        })

        self.assertRedirects(
            response,
            reverse("Tutorias-alumno"),
            fetch_redirect_response=False,
        )
        self.assertEqual(self.client.session["role"], "alumno")

    def test_login_conserva_redireccion_next_sin_enviar_rol(self):
        destination = "/tutorias/in-situ/7/"
        response = self.client.post(
            f'{reverse("login")}?next={destination}',
            {
                "username": self.tutor.email,
                "password": self.password,
                "next": destination,
            },
        )

        self.assertRedirects(response, destination, fetch_redirect_response=False)
        self.assertEqual(self.client.session["role"], "tutor")

    def test_usuario_con_dos_roles_no_puede_iniciar_sesion(self):
        self.tutor.rol = [TUTOR, COORDINADOR]
        self.tutor.save(update_fields=["rol"])

        response = self.client.post(
            reverse("login"),
            {"username": self.tutor.email, "password": self.password},
            follow=True,
        )

        self.assertRedirects(response, reverse("login"))
        self.assertContains(response, "configuración de roles inválida")
        self.assertNotIn("_auth_user_id", self.client.session)


class PushNotificationSettingsTests(TestCase):
    def setUp(self):
        self.user = Usuario.objects.create_user(
            email="push.user@example.com",
            matricula="PUSH001",
            password="password123",
        )
        self.other_user = Usuario.objects.create_user(
            email="push.other@example.com",
            matricula="PUSH002",
            password="password123",
        )
        self.client.force_login(self.user)
        self.subscription_payload = {
            "endpoint": "https://fcm.googleapis.com/fcm/send/device-one",
            "keys": {"auth": "auth-test", "p256dh": "p256dh-test"},
        }
        self.installation_id = "12345678-1234-4234-9234-123456789abc"

    def post_json(self, url, data):
        return self.client.post(
            url,
            data=json.dumps(data),
            content_type="application/json",
        )

    def register_device(self):
        response = self.post_json(reverse("save_webpush_info"), {
            "status_type": "subscribe",
            "subscription": self.subscription_payload,
            "browser": "Chrome",
            "operating_system": "Linux",
            "device_name": "Chrome · Linux",
            "installation_id": self.installation_id,
        })
        self.assertEqual(response.status_code, 201)
        return PushDevice.objects.get(user=self.user)

    def test_pantalla_de_configuracion_conserva_la_ruta_actual(self):
        response = self.client.get(reverse("configuracion_app"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Recibir notificaciones")
        self.assertContains(response, "push_notification_settings.js")

    def test_preferencia_global_no_elimina_dispositivos(self):
        device = self.register_device()

        response = self.post_json(reverse("set_push_preference"), {"enabled": False})

        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertFalse(self.user.notificaciones_habilitadas)
        self.assertTrue(PushDevice.objects.filter(pk=device.pk).exists())

    def test_registro_activa_y_describe_el_dispositivo(self):
        device = self.register_device()

        self.user.refresh_from_db()
        self.assertTrue(self.user.notificaciones_habilitadas)
        self.assertEqual(device.status, PushDevice.Status.ACTIVE)
        self.assertEqual(device.browser, "Chrome")
        self.assertEqual(device.operating_system, "Linux")
        self.assertTrue(
            PushInformation.objects.filter(
                user=self.user,
                subscription=device.subscription,
            ).exists()
        )

    def test_endpoint_se_transfiere_a_la_cuenta_que_lo_activa(self):
        device = self.register_device()
        self.client.force_login(self.other_user)

        response = self.post_json(reverse("save_webpush_info"), {
            "status_type": "subscribe",
            "subscription": self.subscription_payload,
            "browser": "Chrome",
            "operating_system": "Linux",
            "device_name": "Equipo compartido",
            "installation_id": self.installation_id,
        })

        self.assertEqual(response.status_code, 201)
        device.refresh_from_db()
        self.assertEqual(device.user, self.other_user)
        self.assertFalse(
            PushInformation.objects.filter(
                user=self.user,
                subscription=device.subscription,
            ).exists()
        )

    def test_estado_identifica_dispositivo_actual_sin_exponer_endpoint(self):
        device = self.register_device()

        response = self.post_json(reverse("push_notification_state"), {
            "endpoint": self.subscription_payload["endpoint"],
            "installation_id": self.installation_id,
        })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["current_device"]["id"], device.pk)
        self.assertTrue(data["current_device"]["is_current"])
        self.assertTrue(data["current_endpoint_matches"])
        self.assertNotIn("endpoint", data["current_device"])

    def test_estado_asigna_identificador_a_suscripcion_legacy_actual(self):
        device = self.register_device()
        device.installation_id = None
        device.save(update_fields=["installation_id"])

        response = self.post_json(reverse("push_notification_state"), {
            "endpoint": self.subscription_payload["endpoint"],
            "installation_id": self.installation_id,
        })

        self.assertEqual(response.status_code, 200)
        device.refresh_from_db()
        self.assertEqual(str(device.installation_id), self.installation_id)

    def test_nuevo_endpoint_de_la_misma_instalacion_reemplaza_el_anterior(self):
        device = self.register_device()
        original_subscription_id = device.subscription_id
        new_payload = {
            "endpoint": "https://fcm.googleapis.com/fcm/send/device-renewed",
            "keys": {"auth": "new-auth", "p256dh": "new-p256dh"},
        }

        response = self.post_json(reverse("save_webpush_info"), {
            "status_type": "subscribe",
            "subscription": new_payload,
            "browser": "Chrome",
            "operating_system": "Linux",
            "device_name": "Chrome · Linux",
            "installation_id": self.installation_id,
        })

        self.assertEqual(response.status_code, 201)
        device.refresh_from_db()
        self.assertNotEqual(device.subscription_id, original_subscription_id)
        self.assertEqual(device.subscription.endpoint, new_payload["endpoint"])
        self.assertEqual(PushDevice.objects.filter(user=self.user).count(), 1)
        self.assertFalse(
            SubscriptionInfo.objects.filter(pk=original_subscription_id).exists()
        )

    def test_usuario_no_puede_modificar_dispositivo_ajeno(self):
        device = self.register_device()
        self.client.force_login(self.other_user)

        response = self.post_json(
            reverse("set_push_device_status", kwargs={"device_id": device.pk}),
            {"status": PushDevice.Status.PAUSED},
        )

        self.assertEqual(response.status_code, 404)
        device.refresh_from_db()
        self.assertEqual(device.status, PushDevice.Status.ACTIVE)

    def test_dispositivo_puede_pausarse_y_reactivarse_sin_nueva_suscripcion(self):
        device = self.register_device()
        subscription_id = device.subscription_id

        response = self.post_json(
            reverse("set_push_device_status", kwargs={"device_id": device.pk}),
            {"status": PushDevice.Status.PAUSED},
        )
        self.assertEqual(response.status_code, 200)
        device.refresh_from_db()
        self.assertEqual(device.status, PushDevice.Status.PAUSED)

        response = self.post_json(
            reverse("set_push_device_status", kwargs={"device_id": device.pk}),
            {"status": PushDevice.Status.ACTIVE},
        )
        self.assertEqual(response.status_code, 200)
        device.refresh_from_db()
        self.assertEqual(device.status, PushDevice.Status.ACTIVE)
        self.assertEqual(device.subscription_id, subscription_id)

    def test_usuario_puede_cambiar_nombre_de_su_dispositivo(self):
        device = self.register_device()

        response = self.post_json(
            reverse("rename_push_device", kwargs={"device_id": device.pk}),
            {"device_name": "Laptop personal"},
        )

        self.assertEqual(response.status_code, 200)
        device.refresh_from_db()
        self.assertEqual(device.device_name, "Laptop personal")
        self.assertEqual(response.json()["device"]["device_name"], "Laptop personal")

    def test_nombre_de_dispositivo_se_valida(self):
        device = self.register_device()
        url = reverse("rename_push_device", kwargs={"device_id": device.pk})

        self.assertEqual(self.post_json(url, {"device_name": "   "}).status_code, 400)
        self.assertEqual(self.post_json(url, {"device_name": "x" * 151}).status_code, 400)

        device.refresh_from_db()
        self.assertEqual(device.device_name, "Chrome · Linux")

    def test_usuario_no_puede_cambiar_nombre_de_dispositivo_ajeno(self):
        device = self.register_device()
        self.client.force_login(self.other_user)

        response = self.post_json(
            reverse("rename_push_device", kwargs={"device_id": device.pk}),
            {"device_name": "Equipo ajeno"},
        )

        self.assertEqual(response.status_code, 404)
        device.refresh_from_db()
        self.assertEqual(device.device_name, "Chrome · Linux")

    def test_nombre_personalizado_sobrevive_renovacion_de_suscripcion(self):
        device = self.register_device()
        device.device_name = "Laptop personal"
        device.save(update_fields=["device_name"])

        response = self.post_json(reverse("save_webpush_info"), {
            "status_type": "subscribe",
            "subscription": {
                "endpoint": "https://fcm.googleapis.com/fcm/send/device-renamed-renewed",
                "keys": {"auth": "renewed-auth", "p256dh": "renewed-p256dh"},
            },
            "browser": "Chrome",
            "operating_system": "Linux",
            "device_name": "Chrome · Linux",
            "installation_id": self.installation_id,
        })

        self.assertEqual(response.status_code, 201)
        device.refresh_from_db()
        self.assertEqual(device.device_name, "Laptop personal")

    def test_eliminar_dispositivo_borra_suscripcion_tecnica(self):
        device = self.register_device()
        subscription_id = device.subscription_id

        response = self.post_json(
            reverse("delete_push_device", kwargs={"device_id": device.pk}),
            {},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(PushDevice.objects.filter(pk=device.pk).exists())
        self.assertFalse(SubscriptionInfo.objects.filter(pk=subscription_id).exists())

    @patch(
        "Tutorias.signals.handle_push_notifications.send_test_push",
        return_value=(True, "Notificación enviada."),
    )
    def test_prueba_solo_se_envia_al_endpoint_del_dispositivo_actual(self, send_mock):
        device = self.register_device()
        url = reverse("test_push_device", kwargs={"device_id": device.pk})

        wrong_response = self.post_json(url, {"endpoint": "https://example.com/otro"})
        self.assertEqual(wrong_response.status_code, 400)
        send_mock.assert_not_called()

        response = self.post_json(url, {"endpoint": self.subscription_payload["endpoint"]})
        self.assertEqual(response.status_code, 200)
        send_mock.assert_called_once_with(device)

"""Generación de la plantilla XLSX para importar alumnos."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from Usuarios.constants import CARRERAS, ESTADOS_ALUMNO, SEXOS
from Usuarios.services.importacion_alumnos import (
    ENCABEZADOS_IMPORTACION,
    ENCABEZADOS_OPCIONALES,
)


FILA_INICIAL_DATOS = 2
FILA_FINAL_DATOS = 1001
COLOR_PRIMARIO = "F08200"
COLOR_ENCABEZADO = "FFFFFF"

EJEMPLOS_COLUMNAS = {
    "Plan de estudios": ("No", "Biología Molecular"),
    "Matrícula": ("No", "2223028388 (se recomienda formato Texto)"),
    "Correo institucional": ("No", "alumno.ejemplo@cua.uam.mx"),
    "Correo alterno": ("Sí", "alumno.personal@example.com"),
    "Apellido Paterno": ("No", "López"),
    "Apellido Materno": ("Sí", "García"),
    "Nombres": ("No", "María Fernanda"),
    "Núm. económico tutor": ("No", "12345"),
    "Estado académico": ("No", "1"),
    "Sexo": ("No", "Femenino"),
    "Nombre del tutor": ("Sí; columna opcional", "José Antonio Pérez"),
}


def _opciones(opciones):
    return [(codigo, nombre) for codigo, nombre in opciones if codigo != ""]


def _lista_excel(valores):
    """Construye una lista literal para validación de datos de Excel."""
    return '"' + ",".join(str(valor) for valor in valores) + '"'


def _agregar_validacion(hoja, columna, valores, titulo, mensaje):
    validacion = DataValidation(
        type="list",
        formula1=_lista_excel(valores),
        allow_blank=False,
    )
    validacion.error = mensaje
    validacion.errorTitle = "Valor no permitido"
    validacion.prompt = mensaje
    validacion.promptTitle = titulo
    validacion.showErrorMessage = True
    validacion.showInputMessage = True
    hoja.add_data_validation(validacion)
    validacion.add(
        f"{columna}{FILA_INICIAL_DATOS}:{columna}{FILA_FINAL_DATOS}"
    )


def _crear_hoja_alumnos(libro):
    hoja = libro.active
    hoja.title = "Alumnos"
    hoja.freeze_panes = "A2"
    hoja.sheet_view.showGridLines = False
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(ENCABEZADOS_IMPORTACION + ENCABEZADOS_OPCIONALES))}1"

    encabezados = ENCABEZADOS_IMPORTACION + ENCABEZADOS_OPCIONALES
    for indice, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=indice, value=encabezado)
        celda.font = Font(color=COLOR_ENCABEZADO, bold=True)
        celda.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        hoja.column_dimensions[get_column_letter(indice)].width = max(
            18, min(len(encabezado) + 5, 34)
        )
    hoja.row_dimensions[1].height = 30

    matricula_columna = get_column_letter(encabezados.index("Matrícula") + 1)
    for fila in range(FILA_INICIAL_DATOS, FILA_FINAL_DATOS + 1):
        hoja[f"{matricula_columna}{fila}"].number_format = "@"

    posiciones = {nombre: get_column_letter(i + 1) for i, nombre in enumerate(encabezados)}
    _agregar_validacion(
        hoja,
        posiciones["Plan de estudios"],
        [nombre for _, nombre in _opciones(CARRERAS)],
        "Plan de estudios",
        "Seleccione uno de los planes de estudios permitidos.",
    )
    _agregar_validacion(
        hoja,
        posiciones["Estado académico"],
        [codigo for codigo, _ in _opciones(ESTADOS_ALUMNO)],
        "Estado académico",
        "Seleccione un código de estado del 1 al 19.",
    )
    _agregar_validacion(
        hoja,
        posiciones["Sexo"],
        [nombre for _, nombre in _opciones(SEXOS)],
        "Sexo",
        "Seleccione uno de los valores permitidos.",
    )


def _crear_hoja_instrucciones(libro):
    hoja = libro.create_sheet("Instrucciones")
    hoja.sheet_view.showGridLines = False
    hoja["A1"] = "Plantilla para importar alumnos"
    hoja["A1"].font = Font(size=16, bold=True, color=COLOR_PRIMARIO)
    hoja.merge_cells("A1:D1")

    instrucciones = (
        "Capture un alumno por fila en la hoja Alumnos.",
        "No cambie los encabezados ni el orden de la primera hoja.",
        "La matrícula puede ser texto o entero y debe contener 9 o 10 dígitos; se recomienda conservarla como texto.",
        "El tutor debe estar registrado antes de importar el archivo.",
        "Nombre del tutor es opcional y sólo sirve para comprobar el número económico.",
        "El trimestre de ingreso se obtiene automáticamente de la matrícula.",
        "Si existe cualquier error, no se importará ningún alumno.",
    )
    for fila, instruccion in enumerate(instrucciones, start=3):
        hoja.cell(row=fila, column=1, value=f"• {instruccion}")
        hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)

    fila_tabla = len(instrucciones) + 5
    encabezados = ("Columna", "Puede quedar vacía", "Ejemplo", "Observaciones")
    for columna, valor in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila_tabla, column=columna, value=valor)
        celda.font = Font(color=COLOR_ENCABEZADO, bold=True)
        celda.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)

    observaciones = {
        "Matrícula": "Se recomienda conservar formato Texto; también se aceptan enteros.",
        "Núm. económico tutor": "Es el dato utilizado para asignar al tutor.",
        "Estado académico": "Consulte los códigos en la hoja Catálogos.",
        "Nombre del tutor": "No se guarda; genera una advertencia si no coincide.",
    }
    for desplazamiento, encabezado in enumerate(
        ENCABEZADOS_IMPORTACION + ENCABEZADOS_OPCIONALES,
        start=1,
    ):
        puede_vacio, ejemplo = EJEMPLOS_COLUMNAS[encabezado]
        valores = (
            encabezado,
            puede_vacio,
            ejemplo,
            observaciones.get(encabezado, ""),
        )
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila_tabla + desplazamiento, column=columna, value=valor)

    for columna, ancho in zip("ABCD", (32, 22, 42, 58)):
        hoja.column_dimensions[columna].width = ancho
    hoja.freeze_panes = f"A{fila_tabla + 1}"


def _crear_hoja_catalogos(libro):
    hoja = libro.create_sheet("Catálogos")
    hoja.sheet_view.showGridLines = False
    bloques = (
        (1, "Planes de estudios", _opciones(CARRERAS)),
        (4, "Sexo", _opciones(SEXOS)),
        (7, "Estados del alumno", _opciones(ESTADOS_ALUMNO)),
    )
    for columna_inicial, titulo, opciones in bloques:
        titulo_celda = hoja.cell(row=1, column=columna_inicial, value=titulo)
        titulo_celda.font = Font(color=COLOR_ENCABEZADO, bold=True)
        titulo_celda.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        hoja.merge_cells(
            start_row=1,
            start_column=columna_inicial,
            end_row=1,
            end_column=columna_inicial + 1,
        )
        hoja.cell(row=2, column=columna_inicial, value="Código").font = Font(bold=True)
        hoja.cell(row=2, column=columna_inicial + 1, value="Descripción").font = Font(bold=True)
        for fila, (codigo, nombre) in enumerate(opciones, start=3):
            hoja.cell(row=fila, column=columna_inicial, value=codigo)
            hoja.cell(row=fila, column=columna_inicial + 1, value=nombre)

    for columna in ("A", "D", "G"):
        hoja.column_dimensions[columna].width = 12
    for columna in ("B", "E", "H"):
        hoja.column_dimensions[columna].width = 52
    hoja.freeze_panes = "A3"


def generar_plantilla_importacion_alumnos() -> bytes:
    """Devuelve la plantilla actualizada como contenido XLSX en memoria."""
    libro = Workbook()
    _crear_hoja_alumnos(libro)
    _crear_hoja_instrucciones(libro)
    _crear_hoja_catalogos(libro)

    contenido = BytesIO()
    libro.save(contenido)
    return contenido.getvalue()
